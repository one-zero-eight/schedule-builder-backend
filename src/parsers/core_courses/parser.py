import io
import pprint
import re
from collections import defaultdict
from datetime import date, datetime
from itertools import pairwise
from string import ascii_uppercase
from zipfile import ZipFile

import aiohttp
import numpy as np
import pandas as pd
from openpyxl.utils import (
    column_index_from_string,
    coordinate_to_tuple,
    get_column_letter,
)
from yaml import safe_load

from src.domain.dtos.lesson import LessonWithExcelCellsDTO
from src.domain.interfaces.services.parser import ICoursesParser
from src.logging_ import logger
from src.parsers.core_courses.config import core_courses_config as config
from src.parsers.processors.regex import prettify_string
from src.parsers.utils import (
    get_merged_ranges,
    get_sheet_by_id,
    get_sheets,
    split_range_to_xy,
)


class CoreCoursesParser(ICoursesParser):
    """
    Elective parser class
    """

    def __init__(self):
        with open("teachers.yaml", "r", encoding="utf-8") as file:
            self.teachers: list[dict] = safe_load(file)["teachers"]

    def get_clear_dataframes_from_xlsx(
        self, xlsx_file: io.BytesIO, targets: list[config.Target]
    ) -> tuple[dict[str, pd.DataFrame], dict]:
        """
        Get data from xlsx file and return it as a DataFrame with merged
        cells and empty cells in the course row filled by left value.

        :param xlsx_file: xlsx file with data
        :type xlsx_file: io.BytesIO
        :param targets: list of targets to get data from (sheets and ranges)
        :type targets: list[config.Target]

        :return: dataframes with merged cells and empty cells filled
        :rtype: dict[str, pd.DataFrame]
        """
        # ------- Read xlsx file into dataframes -------
        dfs = pd.read_excel(
            xlsx_file, engine="openpyxl", sheet_name=None, header=None
        )
        # ------- Clean up dataframes -------
        dfs = {key.strip(): value for key, value in dfs.items()}
        merged_ranges: dict = {key.strip(): None for key in dfs}
        for target in targets:
            df = None
            for key, value in dfs.items():
                if key.startswith(target.sheet_name):
                    df = value
                    break
            # -------- Fill merged cells with values --------
            merged_ranges[target.sheet_name] = self.merge_cells(
                df, xlsx_file, target.sheet_name
            )
            # -------- Assign excel cell to subject --------
            self.assign_excel_row_and_column_to_subjects(df)
            # -------- Select range --------
            df = self.select_range(df, target.range)
            # -------- Fill empty cells --------
            df = df.replace(r"^\s*$", np.nan, regex=True)
            # -------- Strip, translate and remove trailing spaces --------
            df = df.map(prettify_string)
            # -------- Update dataframe --------
            dfs[target.sheet_name] = df

        return dfs, merged_ranges

    async def get_xlsx_file(self, spreadsheet_id: str) -> io.BytesIO:
        """
        Export xlsx file from Google Sheets and return it as BytesIO object.

        :param spreadsheet_id: id of Google Sheets spreadsheet
        :return: xlsx file as BytesIO object
        """
        # ------- Get data from Google Sheets -------
        # ------- Create url for export -------
        spreadsheet_url = (
            f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        )
        export_url = spreadsheet_url + "/export?format=xlsx"
        # ------- Export xlsx file -------
        async with aiohttp.ClientSession() as client:
            async with client.get(export_url) as response:
                return io.BytesIO(await response.read())

    def check_value_is_time(self, value: str) -> bool:
        value = value.split("-")
        if len(value) != 2:
            return False
        for part in value:
            part = part.strip().split(":")
            if len(part) != 2:
                return False
            if not part[0].isnumeric() or not part[1].isnumeric():
                return False
        return True

    def check_is_weekday(self, value: str) -> bool:
        return value in [
            "MONDAY",
            "TUESDAY",
            "WEDNESDAY",
            "THURSDAY",
            "FRIDAY",
            "SATURDAY",
            "SUNDAY",
        ]

    def coords_to_excel_coords(self, i: int, j: int) -> str:
        excel_row = i + 1
        excel_col = []
        tmp = j
        while tmp > 25:
            excel_col.append(ascii_uppercase[tmp // 26 - 1])
            tmp //= 26
        excel_col.append(ascii_uppercase[j % 26])
        excel_col = "".join(excel_col)
        return f"{excel_col}{excel_row}"

    def assign_excel_row_and_column_to_subjects(
        self, df: pd.DataFrame
    ) -> None:
        used_cells: set[tuple[int, int]] = set()
        for i in range(3, len(df.values)):
            for j in range(1, len(df.values[i])):
                value = str(df.iloc[i, j]).strip()
                if (
                    not value
                    or value == "nan"
                    or self.check_is_weekday(value)
                    or self.check_value_is_time(value)
                ):
                    continue
                if (i, j) in used_cells:
                    continue
                df.iloc[i, j] = (
                    f"{df.iloc[i, j]}${self.coords_to_excel_coords(i, j)}"
                )
                for x in range(i, i + 3):
                    used_cells.add((x, j))

    def merge_cells(
        self, df: pd.DataFrame, xlsx: io.BytesIO, target_sheet_name: str
    ):
        """
        Merge cells in dataframe

        :param df: Dataframe to process
        :param xlsx: xlsx file with data
        :param target_sheet_name: sheet to process
        """
        xlsx.seek(0)
        xlsx_zipfile = ZipFile(xlsx)
        sheets = get_sheets(xlsx_zipfile)
        target_sheet_id = None
        for sheet_id, sheet_name in sheets.items():
            if target_sheet_name in sheet_name:
                target_sheet_id = sheet_id
                break
        sheet = get_sheet_by_id(xlsx_zipfile, target_sheet_id)
        merged_ranges = get_merged_ranges(sheet)

        # ------- Merge cells -------
        for merged_range in merged_ranges:
            (start_row, start_col), (end_row, end_col) = split_range_to_xy(
                merged_range
            )
            # get value from top left cell
            value = df.iloc[start_row, start_col]
            # fill merged cells with value
            df.iloc[start_row : end_row + 1, start_col : end_col + 1] = value
        return merged_ranges

    def select_range(
        self, df: pd.DataFrame, target_range: str
    ) -> pd.DataFrame:
        """
        Select range from dataframe

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param target_range: range to select
        :type target_range: str
        :return: selected range
        :rtype: pd.DataFrame
        """
        (start_row, start_col), (end_row, end_col) = split_range_to_xy(
            target_range
        )
        return df.iloc[
            start_row : end_row + 1,
            start_col : end_col + 1,
        ]

    def set_weekday_and_time_as_index(
        self, df: pd.DataFrame, column: int = 0
    ) -> pd.DataFrame:
        """
        Set time column as index and process it to datetime format

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param column: column to set as index, defaults to 0
        :type column: int, optional
        """

        # get column view and iterate over it
        df_column = df.iloc[:, column]
        df_column: pd.Series
        # drop column
        df.drop(df.columns[column], axis=1, inplace=True)
        # fill nan values with previous value
        df_column.ffill(inplace=True)

        # ----- Process weekday ------ #
        # get indexes of weekdays
        weekdays_indexes = [
            i
            for i, cell in enumerate(df_column.values)
            if cell in config.WEEKDAYS
        ]

        # create index mapping for weekdays [None, None, "MONDAY", "MONDAY", ...]
        index_mapping = pd.Series(index=df_column.index, dtype=object)
        last_index = len(df_column)
        for start, end in pairwise(weekdays_indexes + [last_index]):
            index_mapping.iloc[start] = "delete"
            index_mapping.iloc[start + 1 : end] = df_column[start]

        # ----- Process time ------ #
        # matched r"\d{1,2}:\d{2}-\d{1,2}:\d{2}" regex

        matched = df_column[
            df_column.str.match(r"\d{1,2}:\d{2}-\d{1,2}:\d{2}")
        ]

        for i, cell in matched.items():
            # "9:00-10:30" -> datetime.time(9, 0), datetime.time(10, 30)
            start, end = cell.split("-")
            df_column.loc[i] = (
                datetime.strptime(start, "%H:%M").time(),
                datetime.strptime(end, "%H:%M").time(),
            )

        # create multiindex from index mapping and time column
        multiindex = pd.MultiIndex.from_arrays(
            [index_mapping, df_column], names=["weekday", "time"]
        )
        # set multiindex as index
        df.set_index(multiindex, inplace=True)
        # drop rows with weekday
        df.drop("delete", inplace=True, level=0)
        return df

    def set_course_and_group_as_header(
        self, df: pd.DataFrame, rows: tuple = (0, 1)
    ) -> pd.DataFrame:
        """
        Set course and group as header

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param rows: row to set as columns, defaults to (0, 1)
        :type rows: tuple, optional
        """
        # ------- Set course and group as header -------
        # get rows with course and group
        df_header = df.iloc[rows[0] : rows[1] + 1]
        # drop rows with course and group
        df.drop(list(rows), inplace=True)
        df.reset_index(drop=True, inplace=True)
        # fill nan values with previous value
        with pd.option_context("future.no_silent_downcasting", True):
            df_header = df_header.ffill(axis=1)
        multiindex = pd.MultiIndex.from_arrays(
            df_header.values, names=["course", "group"]
        )
        df.columns = multiindex
        return df

    def split_df_by_courses(
        self, df: pd.DataFrame, time_columns: list[int]
    ) -> list[pd.DataFrame]:
        """
        Split dataframe by "Week *" rows

        :param time_columns: list of columns(pd) with time and weekday
        :param df: dataframe to split
        :type df: pd.DataFrame
        :return: list of dataframes with locators
        :rtype: list[pd.DataFrame, ExcelToPandasLocator]
        """

        time_columns_indexes = time_columns

        # split dataframe by found columns
        _, max_y = df.shape
        split_indexes = time_columns_indexes + [max_y]

        split_dfs = []

        for _, (start, end) in enumerate(pairwise(split_indexes)):
            split_df = df.iloc[:, start:end].copy()
            split_dfs.append(split_df)
        return split_dfs

    def identify_room(self, location: str) -> str:
        location = location.strip().rstrip()
        if "ONLINE" in location:
            return location
        guess = location.split()[0]
        if not guess.isnumeric():
            return location
        return guess

    def _parse_schedule_string(
        self, s: str
    ) -> list[tuple[str, date | None, date | None]]:
        # удаляю время
        s = re.sub(r"STARTS AT \d{1,2}:\d{2}", "", s)
        s = s.replace("ТОЛЬКО НА", "ON")

        # извлекаю всё из скобок
        bracket_parts = re.findall(r"\((.*?)\)", s)
        s = re.sub(r"\(.*?\)", "", s)

        # разбиваю главную часть на другие части, если они есть
        parts = [p.strip() for p in re.split(r"\|", s) if p.strip()]
        parts.extend(bracket_parts)

        result = []
        for part in parts:
            part = part.strip()
            if not part:
                continue

            # отдельно для online
            if part.startswith("ONLINE"):
                content = part[6:].strip()
                if content.startswith("ON"):
                    content = content[2:].strip()
                    dates = re.findall(r"\d{1,2}/\d{2}", content)
                    if dates:
                        for d in dates:
                            result.append(("ONLINE", date(d), None))
                        continue
                elif content.startswith("EXCEPT"):
                    content = content[6:].strip()
                    dates = re.findall(r"\d{1,2}/\d{2}", content)
                    if dates:
                        for d in dates:
                            result.append(("ONLINE", None, date(d)))
                    continue
                result.append(("ONLINE", None, None))
                continue

            room_match = re.match(r"^[\d /]+", part)
            if not room_match:
                continue

            room_str = room_match.group(0)
            rooms = room_str.split("/")
            rest = part[len(room_str) :].strip()

            on_dates = []
            except_dates = []
            rest = rest.replace("STARTS ON", "ON")

            if "ON" in rest:
                rest_after_on = rest[rest.find("ON") + 2 :].strip()
                dates = re.findall(r"\d{1,2}/\d{2}", rest_after_on)
                if dates:
                    on_dates = dates

            if "EXCEPT" in rest:
                dates_str = rest[: rest.find("EXCEPT")].strip()
                except_str = rest[rest.find("EXCEPT") + 6 :].strip()

                dates = re.findall(r"\d{1,2}/\d{2}", dates_str)
                except_dates_list = re.findall(r"\d{1,2}/\d{2}", except_str)

                if dates:
                    on_dates = dates
                if except_dates_list:
                    except_dates = except_dates_list

            if len(on_dates) == 0 and len(except_dates) == 0:
                for room in rooms:
                    result.append((room.strip(), None, None))

            for room in rooms:
                # ON и EXCEPT разом не должны выпасть
                for d in on_dates:
                    result.append((room.strip(), date(d), None))
                for d in except_dates:
                    result.append((room.strip(), None, date(d)))

        return result

    def parse_schedule_string(
        self, s: str
    ) -> list[tuple[str, date | None, date | None]]:
        try:
            return self._parse_schedule_string(s)
        except Exception as e:
            logger.error(f"Не удалось спарсить {s}, error: {e}")
            return [(s, None, None)]

    async def get_all_timeslots(
        self, spreadsheet_id: str
    ) -> list[LessonWithExcelCellsDTO]:
        xlsx = await self.get_xlsx_file(spreadsheet_id=spreadsheet_id)
        dfs, dfs_merged_ranges = self.get_clear_dataframes_from_xlsx(
            xlsx_file=xlsx, targets=config.TARGETS
        )
        lessons = []

        for target in config.TARGETS:
            # find dataframe from dfs
            sheet_df = dfs[target.sheet_name]
            merged_ranges = [
                split_range_to_xy(merged_range)
                for merged_range in dfs_merged_ranges[target.sheet_name]
            ]
            # [(start_row, start_col), (end_row, end_col)]
            time_columns_index = [
                column_index_from_string(col) - 1
                for col in target.time_columns
            ]
            by_courses = self.split_df_by_courses(sheet_df, time_columns_index)
            for course_df in by_courses:
                course_lessons = []
                # -------- Set course and group as header; weekday and timeslot as index --------
                self.set_course_and_group_as_header(course_df)
                self.set_weekday_and_time_as_index(course_df)

                for timeslot, timeslot_df in course_df.groupby(
                    level=[0, 1], sort=False
                ):
                    weekday, (start_time, end_time) = timeslot
                    for column, cell_values_series in timeslot_df.items():
                        cell_values_series: pd.Series
                        year, group = column
                        if pd.isna(cell_values_series).all():
                            continue
                        else:
                            subject, teacher, location = (
                                cell_values_series.values
                            )
                            subject = subject.rsplit("$", maxsplit=1)
                            subject_name, cell = subject
                            location = self.parse_schedule_string(
                                str(location)
                            )
                            group = group.split()
                            group_name = group[0]
                            if len(group) > 1:
                                students_number = int(group[1][1:-1])
                            else:
                                students_number = 1
                        for loc, on_, except_ in location:
                            course_lessons.append(
                                LessonWithExcelCellsDTO(
                                    weekday=weekday,
                                    start_time=start_time,
                                    end_time=end_time,
                                    group_name=group_name,
                                    teacher=teacher,
                                    teacher_email=next(
                                        (
                                            t.get("email", "")
                                            for t in self.teachers
                                            if t.get("name") == teacher
                                        ),
                                        "",
                                    ),
                                    room=loc,
                                    lesson_name=subject_name,
                                    students_number=students_number,
                                    excel_range=cell,
                                    date_on=on_,
                                    date_except=except_,
                                )
                            )

                merged_registry = defaultdict(
                    list
                )  # merged range index X list of lessons
                non_merged = []
                for lesson in course_lessons:
                    cell_row, cell_col = coordinate_to_tuple(
                        lesson.excel_range
                    )
                    cell_row -= 1
                    cell_col -= 1
                    for i, (
                        (start_row, start_col),
                        (end_row, end_col),
                    ) in enumerate(merged_ranges):
                        if (start_row <= cell_row <= end_row) and (
                            start_col <= cell_col <= end_col
                        ):
                            merged_registry[i].append(lesson)
                            break
                    else:  # non_merged
                        non_merged.append(lesson)

                merged = []
                for _lessons in merged_registry.values():
                    _lessons: list[LessonWithExcelCellsDTO]
                    groups = []
                    students_number = []
                    excel_ranges = []
                    lesson1 = _lessons[0]
                    for lesson in _lessons:
                        if isinstance(lesson.group_name, list):
                            groups.extend(lesson.group_name)
                        elif lesson.group_name:
                            groups.append(lesson.group_name)
                        if lesson.students_number:
                            students_number.append(lesson.students_number)
                        if lesson.excel_range:
                            excel_ranges.append(lesson.excel_range)
                    lesson1.group_name = groups
                    lesson1.students_number = (
                        sum(students_number)
                        if students_number
                        else lesson1.students_number
                    )
                    rows = {
                        "".join(filter(str.isdigit, c)) for c in excel_ranges
                    }
                    if len(rows) != 1:
                        raise ValueError("Cells are not on the same row")
                    row = rows.pop()
                    col_numbers = sorted(
                        column_index_from_string(
                            "".join(filter(str.isalpha, c))
                        )
                        for c in excel_ranges
                    )
                    first = get_column_letter(col_numbers[0]) + row
                    last = get_column_letter(col_numbers[-1]) + row
                    lesson1.excel_range = f"{first}:{last}"
                    merged.append(lesson1)

                logger.info(f"Merged lessons: {pprint.pformat(merged)}")
                lessons.extend(merged)
                logger.info(
                    f"Non-merged lessons: {pprint.pformat(non_merged)}"
                )
                lessons.extend(non_merged)
        return lessons
