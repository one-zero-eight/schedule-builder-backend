import datetime
import io
import re
from collections import defaultdict
from itertools import pairwise
from string import ascii_uppercase

import httpx
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import (
    column_index_from_string,
    coordinate_to_tuple,
    get_column_letter,
)

from src.logging_ import logger
from src.modules.options.repository import options_repository
from src.modules.parsers.core_courses.location_parser import Item, parse_location_string
from src.modules.parsers.processors.regex import prettify_string
from src.modules.parsers.schemas import Lesson
from src.modules.parsers.utils import sanitize_sheet_name

WEEKDAYS = [
    "MONDAY",
    "TUESDAY",
    "WEDNESDAY",
    "THURSDAY",
    "FRIDAY",
    "SATURDAY",
    "SUNDAY",
]


def nearest_weekday(date: datetime.date, day: int | str) -> datetime.date:
    """
    Returns the date of the next given weekday after
    the given date. For example, the date of next Monday.

    :param date: date to start from
    :type date: datetime.date
    :param day: weekday to find (0 is Monday, 6 is Sunday)
    :type day: int
    :return: date of the next given weekday
    :rtype: datetime.date
    """
    if isinstance(day, str):
        day = ["mo", "tu", "we", "th", "fr", "sa", "su"].index(day[:2].lower())

    days = (day - date.weekday() + 7) % 7
    return date + datetime.timedelta(days=days)


class CoreCoursesParser:
    """
    Elective parser class
    """

    def get_clear_dataframes_from_xlsx(
        self, xlsx_file: io.BytesIO, target_sheet_names: list[str]
    ) -> tuple[dict[str, pd.DataFrame], dict]:
        """
        Get data from xlsx file and return it as a DataFrame with merged
        cells and empty cells in the course row filled by left value.

        :param xlsx_file: xlsx file with data
        :type xlsx_file: io.BytesIO
        :param targets: list of targets to get data from (sheets and ranges)
        :type targets: list[config.Target]

        :return: dataframes with merged cells and empty cells filled
        :rtype: dict[str, pd.DataFrame]
        """
        # ------- Read xlsx file into dataframes -------
        dfs = pd.read_excel(xlsx_file, engine="openpyxl", sheet_name=None, header=None)
        # ------- Clean up dataframes -------
        merged_ranges: dict[str, list[tuple[int, int, int, int]]] = defaultdict(list)
        for target_sheet_name in target_sheet_names:
            df = dfs[target_sheet_name]
            # -------- Select range --------
            (min_row, min_col, max_row, max_col) = self.auto_detect_range(df, xlsx_file, target_sheet_name)
            df = df.iloc[min_row : max_row + 1, min_col : max_col + 1]
            # -------- Fill merged cells with values --------
            merged_ranges[target_sheet_name] = self.merge_cells(df, xlsx_file, target_sheet_name)
            # -------- Assign excel cell to subject --------
            self.assign_excel_row_and_column_to_subjects(df)
            # -------- Fill empty cells --------
            df = df.replace(r"^\s*$", np.nan, regex=True)
            # -------- Strip, translate and remove trailing spaces --------
            df = df.map(prettify_string)
            # -------- Update dataframe --------
            dfs[target_sheet_name] = df

        return dfs, merged_ranges

    async def get_xlsx_file(self, spreadsheet_id: str) -> io.BytesIO:
        """
        Export xlsx file from Google Sheets and return it as BytesIO object.

        :param spreadsheet_id: id of Google Sheets spreadsheet
        :return: xlsx file as BytesIO object
        """
        # ------- Get data from Google Sheets -------
        # ------- Create url for export -------
        spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        export_url = spreadsheet_url + "/export?format=xlsx"
        # ------- Export xlsx file -------
        async with httpx.AsyncClient() as client:
            response = await client.get(export_url, follow_redirects=True)
            response.raise_for_status()
            return io.BytesIO(response.content)

    def check_value_is_time(self, value: str) -> bool:
        value = value.split("-")
        if len(value) != 2:
            return False
        for part in value:
            part = part.strip().split(":")  # noqa: PLW2901
            if len(part) != 2:
                return False
            if not part[0].isnumeric() or not part[1].isnumeric():
                return False
        return True

    def check_is_weekday(self, value: str) -> bool:
        return value in WEEKDAYS

    def coords_to_excel_coords(self, i: int, j: int) -> str:
        excel_row = i + 1
        excel_col = []
        tmp = j
        while tmp > 25:
            excel_col.append(ascii_uppercase[tmp // 26 - 1])
            tmp //= 26
        excel_col.append(ascii_uppercase[j % 26])
        excel_col = "".join(excel_col)
        return f"{excel_col}{excel_row}"

    def assign_excel_row_and_column_to_subjects(self, df: pd.DataFrame) -> None:
        used_cells: set[tuple[int, int]] = set()
        for i in range(3, len(df.values)):
            for j in range(1, len(df.values[i])):
                value = str(df.iloc[i, j]).strip()
                if not value or value == "nan" or self.check_is_weekday(value) or self.check_value_is_time(value):
                    continue
                if (i, j) in used_cells:
                    continue
                df.iloc[i, j] = f"{df.iloc[i, j]}${self.coords_to_excel_coords(i, j)}"
                for x in range(i, i + 3):
                    used_cells.add((x, j))

    def merge_cells(
        self, df: pd.DataFrame, xlsx: io.BytesIO, target_sheet_name: str
    ) -> list[tuple[int, int, int, int]]:
        """
        Merge cells in dataframe

        :param df: Dataframe to process
        :param xlsx: xlsx file with data
        :param target_sheet_name: sheet to process
        :return: list of merged ranges: (min_row, min_col, max_row, max_col)
        """
        xlsx.seek(0)
        ws = openpyxl.load_workbook(xlsx)
        sheet = ws[target_sheet_name]
        merged_ranges = []
        # ------- Merge cells -------
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            min_col = min_col - 1
            min_row = min_row - 1
            max_col = max_col - 1
            max_row = max_row - 1
            value = df.iloc[min_row, min_col]

            df.iloc[min_row : max_row + 1, min_col : max_col + 1] = value
            merged_ranges.append((min_row, min_col, max_row, max_col))
        return merged_ranges

    def set_weekday_and_time_as_index(self, df: pd.DataFrame, column: int = 0) -> pd.DataFrame:
        """
        Set time column as index and process it to datetime format

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param column: column to set as index, defaults to 0
        :type column: int, optional
        """

        # get column view and iterate over it
        df_column = df.iloc[:, column]
        df_column: pd.Series
        # drop column
        df.drop(df.columns[column], axis=1, inplace=True)

        # ----- Process weekday ------ #
        # get indexes of weekdays
        weekdays_indexes = [i for i, cell in enumerate(df_column.values) if cell in WEEKDAYS]

        # create index mapping for weekdays [None, None, "MONDAY", "MONDAY", ...]
        index_mapping = pd.Series(index=df_column.index, dtype=object)
        last_index = len(df_column)
        for start, end in pairwise(weekdays_indexes + [last_index]):
            index_mapping.iloc[start] = "delete"
            index_mapping.iloc[start + 1 : end] = df_column[start]

        # ----- Process time ------ #
        # matched r"\d{1,2}:\d{2}-\d{1,2}:\d{2}" regex
        mask = df_column.astype(str).str.match(r"\d{1,2}:\d{2}-\d{1,2}:\d{2}") & df_column.notna()
        matched = df_column[mask]
        for i, cell in matched.items():
            # "9:00-10:30" -> datetime.time(9, 0), datetime.time(10, 30)
            start, end = cell.split("-")
            df_column.loc[i] = (
                datetime.datetime.strptime(start, "%H:%M").time(),
                datetime.datetime.strptime(end, "%H:%M").time(),
            )

        # create multiindex from index mapping and time column
        multiindex = pd.MultiIndex.from_arrays([index_mapping, df_column], names=["weekday", "time"])
        # set multiindex as index
        df.set_index(multiindex, inplace=True)
        # drop rows with weekday
        df.drop("delete", inplace=True, level=0)
        return df

    def set_course_and_group_as_header(self, df: pd.DataFrame, rows: tuple = (0, 1)) -> pd.DataFrame:
        """
        Set course and group as header

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param rows: row to set as columns, defaults to (0, 1)
        :type rows: tuple, optional
        """
        # ------- Set course and group as header -------
        # get rows with course and group
        df_header = df.iloc[rows[0] : rows[1] + 1]
        # drop rows with course and group
        df.drop(list(rows), inplace=True)
        df.reset_index(drop=True, inplace=True)
        # fill nan values with previous value
        with pd.option_context("future.no_silent_downcasting", True):
            df_header = df_header.ffill(axis=1)
        multiindex = pd.MultiIndex.from_arrays(df_header.values, names=["course", "group"])
        df.columns = multiindex
        return df

    def split_df_by_courses(self, df: pd.DataFrame, time_columns: list[int]) -> list[pd.DataFrame]:
        """
        Split dataframe by "Week *" rows

        :param time_columns: list of columns(pd) with time and weekday
        :param df: dataframe to split
        :type df: pd.DataFrame
        :return: list of dataframes with locators
        :rtype: list[pd.DataFrame, ExcelToPandasLocator]
        """

        time_columns_indexes = time_columns

        # split dataframe by found columns
        _, max_y = df.shape
        split_indexes = time_columns_indexes + [max_y]

        split_dfs = []

        for _, (start, end) in enumerate(pairwise(split_indexes)):
            split_df = df.iloc[:, start:end].copy()
            split_dfs.append(split_df)
        return split_dfs

    def auto_detect_range(
        self, sheet_df: pd.DataFrame, xlsx_file: io.BytesIO, sheet_name: str
    ) -> tuple[int, int, int, int]:
        """
        :return: tuple of (min_row, min_col, max_row, max_col)
        """
        time_columns_index = self.get_time_columns(sheet_df)
        logger.info(f"Time columns: {[get_column_letter(col + 1) for col in time_columns_index]}")
        # -------- Get rightmost column index --------
        rightmost_column_index = self.get_rightmost_column_index(xlsx_file, sheet_name, time_columns_index)
        logger.info(f"Rightmost column index: {get_column_letter(rightmost_column_index + 1)}")
        last_row_index = self.get_last_row_index(xlsx_file, sheet_name)
        target_range = f"A1:{get_column_letter(rightmost_column_index + 1)}{last_row_index}"
        logger.info(f"Target range: {target_range}")
        return (0, 0, last_row_index, rightmost_column_index)

    def get_time_columns(self, sheet_df: pd.DataFrame) -> list[int]:
        # find columns where presents "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"
        time_columns = []
        for column in sheet_df.columns:
            values_in_column = sheet_df[column].values
            if all(weekday in values_in_column for weekday in WEEKDAYS[:-1]):
                time_columns.append(column)
        return time_columns

    def get_rightmost_column_index(self, xlsx_file: io.BytesIO, sheet_name: str, time_columns: list[int]) -> int:
        # Column after time columns that has no borders formatting

        wb = openpyxl.load_workbook(xlsx_file)
        sheet = wb[sheet_name]
        last_time_column = time_columns[-1]

        next_column = last_time_column + 1
        cell = sheet.cell(row=1, column=next_column + 1)

        # Check if cell has no borders
        has_no_border = (
            (cell.border is None or cell.border.right is None or cell.border.right.style is None)
            and (cell.border is None or cell.border.top is None or cell.border.top.style is None)
            and (cell.border is None or cell.border.bottom is None or cell.border.bottom.style is None)
        )

        if has_no_border:
            return next_column - 1
        else:
            # Continue searching for column with no border
            for col in range(next_column + 1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col + 1)
                if (
                    (cell.border is None or cell.border.right is None or cell.border.right.style is None)
                    and (cell.border is None or cell.border.top is None or cell.border.top.style is None)
                    and (cell.border is None or cell.border.bottom is None or cell.border.bottom.style is None)
                ):
                    return col - 1
            return next_column  # fallback

    def get_last_row_index(self, xlsx_file: io.BytesIO, sheet_name: str) -> int:
        wb = openpyxl.load_workbook(xlsx_file)
        sheet = wb[sheet_name]
        return sheet.max_row

    async def get_all_lessons(self, spreadsheet_id: str, target_sheet_names: list[str]) -> list[Lesson]:
        original_target_sheet_names = target_sheet_names
        sanitized_sheet_names = [sanitize_sheet_name(target_sheet_name) for target_sheet_name in target_sheet_names]
        semester = options_repository.get_semester()
        assert semester is not None, "Semester is not set"

        xlsx = await self.get_xlsx_file(spreadsheet_id=spreadsheet_id)
        dfs, dfs_merged_ranges = self.get_clear_dataframes_from_xlsx(
            xlsx_file=xlsx, target_sheet_names=sanitized_sheet_names
        )
        lessons = []

        for target_sheet_name, original_target_sheet_name in zip(sanitized_sheet_names, original_target_sheet_names):
            # find dataframe from dfs
            if target_sheet_name not in dfs:
                logger.warning(f"Sheet {target_sheet_name} not found in xlsx file")
                continue
            sheet_df = dfs[target_sheet_name]
            merged_ranges = dfs_merged_ranges[target_sheet_name]
            # [(start_row, start_col), (end_row, end_col)]

            time_columns_index = self.get_time_columns(sheet_df)
            logger.info(f"Sheet Time columns: {[get_column_letter(col + 1) for col in time_columns_index]}")
            rightmost_column_index = self.get_rightmost_column_index(xlsx, target_sheet_name, time_columns_index)
            logger.info(f"Rightmost column index: {get_column_letter(rightmost_column_index + 1)}")
            by_courses = self.split_df_by_courses(sheet_df, time_columns_index)
            for course_df in by_courses:
                course_lessons = []
                # -------- Set course and group as header; weekday and timeslot as index --------
                self.set_course_and_group_as_header(course_df)
                self.set_weekday_and_time_as_index(course_df)

                for timeslot, timeslot_df in course_df.groupby(level=[0, 1], sort=False):
                    weekday, (start_time, end_time) = timeslot
                    for column, cell_values_series in timeslot_df.items():
                        cell_values_series: pd.Series
                        course_name, group = column
                        group_name: str | tuple[str, ...]
                        students_number: int
                        if isinstance(group, str):
                            # find student number in B24-CSE-02 (29) format
                            if m := re.search(r"\((\d+)\)", group):
                                students_number = int(m.group(1))
                                group_name = re.sub(r"\(\d+\)", "", group).strip()
                            else:
                                students_number = 0
                                group_name = group.strip()
                        else:
                            students_number = 0
                            group_name = tuple()

                        if pd.isna(cell_values_series).all():
                            continue
                        else:
                            try:
                                subject, teacher, location = cell_values_series.values
                            except ValueError:
                                logger.warning(
                                    f"Cell values: {cell_values_series.values} for column={column} and timeslot={timeslot}"
                                )
                                raise
                            if not isinstance(subject, str):
                                logger.warning(f"Subject {subject} not found in xlsx file")
                                logger.warning(f"Cell values: {cell_values_series.values}")

                            subject = subject.rsplit("$", maxsplit=1)
                            if len(subject) == 1:
                                logger.warning(f"Subject {subject} not found in xlsx file")
                            subject_name, cell = subject

                            if pd.isna(teacher):
                                teacher = None
                            if pd.isna(location):
                                location = None
                            for v in (subject_name, course_name, group_name):
                                if (
                                    v == "Elective courses on Physical Education"
                                    or location == "Elective course on Physical Education"
                                ):
                                    subject_name = "Elective courses on Physical Education"
                                    location = None
                                    teacher = None
                                    break

                            if location:
                                location_str = str(location)
                                location_item = parse_location_string(location_str)
                                if not location_item:
                                    logger.warning(f"Location {location_str} cannot be parsed")
                            else:
                                location_str = None
                                location_item = None

                            if location_item is None:
                                course_lessons.append(
                                    Lesson(
                                        lesson_name=subject_name,
                                        weekday=weekday,
                                        start_time=start_time,
                                        end_time=end_time,
                                        course_name=course_name,
                                        group_name=group_name,
                                        teacher=teacher,
                                        room=location_str,
                                        date_on=None,
                                        date_except=None,
                                        students_number=students_number,
                                        excel_sheet_name=original_target_sheet_name,
                                        excel_range=cell,
                                    )
                                )
                            else:
                                starts = location_item.starts_from or semester.start_date

                                def convert_weeks_on_to_only_on(item: Item):
                                    if item.on_weeks:
                                        on = []
                                        for week in item.on_weeks:
                                            on_date = nearest_weekday(starts, weekday) + datetime.timedelta(
                                                weeks=week - 1
                                            )
                                            on.append(on_date)
                                        if item.on:
                                            item.on.extend(on)
                                        elif on:
                                            item.on = on
                                    if item.on:
                                        item.on = sorted(set(item.on))

                                lesson_start_time = start_time
                                lesson_end_time = end_time

                                if location_item.starts_at:
                                    _start_time = datetime.datetime.combine(starts, start_time)
                                    _end_time = datetime.datetime.combine(starts, end_time)
                                    duration = _end_time - _start_time
                                    lesson_start_time = location_item.starts_at
                                    lesson_end_time = (
                                        datetime.datetime.combine(starts, lesson_start_time) + duration
                                    ).time()
                                if location_item.till:
                                    lesson_end_time = location_item.till

                                convert_weeks_on_to_only_on(location_item)

                                main_lesson = Lesson(
                                    lesson_name=subject_name,
                                    weekday=weekday,
                                    start_time=lesson_start_time,
                                    end_time=lesson_end_time,
                                    course_name=course_name,
                                    group_name=group_name,
                                    teacher=teacher,
                                    room=location_item.location or location_str,
                                    date_on=location_item.on,
                                    date_except=location_item.except_,
                                    date_from=location_item.starts_from,
                                    students_number=students_number,
                                    excel_sheet_name=original_target_sheet_name,
                                    excel_range=cell,
                                )

                                course_lessons.append(main_lesson)

                                nested_on: list[Item] = []
                                extra_nested: list[Item] = []
                                if location_item.NEST:
                                    for item in location_item.NEST:
                                        convert_weeks_on_to_only_on(item)
                                        if item.on:
                                            nested_on.append(item)
                                        else:
                                            logger.info(f"Root Item: {location_item}, {item}")
                                            extra_nested.append(item)

                                if extra_nested:  # TODO: Handle '421 (316 FROM 31/10)' case
                                    logger.warning(f"Extra nested is not implemented yet\nItem({location_item})")

                                for item in nested_on:
                                    assert item.on, f"Item {item} has no on"
                                    if item.location:
                                        main_lesson.date_except = (main_lesson.date_except or []) + item.on

                                    nested_lesson = main_lesson.model_copy()
                                    nested_lesson.date_on = item.on
                                    nested_lesson.room = item.location or main_lesson.room
                                    nested_lesson.start_time = item.starts_at or main_lesson.start_time
                                    nested_lesson.end_time = item.till or main_lesson.end_time
                                    course_lessons.append(nested_lesson)

                merged_registry = defaultdict(list)  # merged range index X list of lessons
                non_merged = []
                for lesson in course_lessons:
                    cell_row, cell_col = coordinate_to_tuple(lesson.excel_range)
                    cell_row -= 1
                    cell_col -= 1
                    for i, (min_row, min_col, max_row, max_col) in enumerate(merged_ranges):
                        if (min_row <= cell_row <= max_row) and (min_col <= cell_col <= max_col):
                            merged_registry[i].append(lesson)
                            break
                    else:  # non_merged
                        non_merged.append(lesson)

                merged = []
                for _lessons in merged_registry.values():
                    _lessons: list[Lesson]
                    groups = []
                    merged_students_number = []
                    excel_ranges = []
                    lesson1 = _lessons[0]
                    for lesson in _lessons:
                        if lesson.group_name:
                            if lesson.group_name not in groups and lesson.students_number:
                                merged_students_number.append(lesson.students_number)
                            if lesson.group_name not in groups and lesson.excel_range:
                                excel_ranges.append(lesson.excel_range)
                        else:
                            if lesson.students_number:
                                merged_students_number.append(lesson.students_number)
                            if lesson.excel_range:
                                excel_ranges.append(lesson.excel_range)

                        if isinstance(lesson.group_name, list | tuple):
                            groups.extend(lesson.group_name)
                        elif lesson.group_name:
                            groups.append(lesson.group_name)

                        assert lesson.excel_sheet_name == lesson1.excel_sheet_name, (
                            "All lessons should be in the same sheet"
                        )

                    logger.info(groups)
                    logger.info(excel_ranges)
                    logger.info(merged_students_number)

                    lesson1.group_name = tuple(groups)
                    lesson1.students_number = (
                        sum(merged_students_number) if merged_students_number else lesson1.students_number
                    )
                    rows = {"".join(filter(str.isdigit, c)) for c in excel_ranges}
                    if len(rows) != 1:
                        raise ValueError(f"Cells are not on the same row: {excel_ranges}")
                    row = rows.pop()
                    col_numbers = sorted(
                        column_index_from_string("".join(filter(str.isalpha, c))) for c in excel_ranges
                    )
                    first = get_column_letter(col_numbers[0]) + row
                    last = get_column_letter(col_numbers[-1]) + row
                    lesson1.excel_range = f"{first}:{last}"
                    merged.append(lesson1)

                lessons.extend(merged)
                lessons.extend(non_merged)
        return lessons
