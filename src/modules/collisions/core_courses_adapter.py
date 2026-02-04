import datetime
from collections import defaultdict
from collections.abc import Generator

import pandas as pd
from openpyxl.utils import coordinate_to_tuple, get_column_letter

from src.core_courses.cell_to_event import CoreCourseEvent, convert_cell_to_event
from src.core_courses.config import CoreCoursesConfig, Target
from src.core_courses.location_parser import Item
from src.core_courses.parser import CoreCourseCell, CoreCoursesParser
from src.logging_ import logger
from src.utils import WEEKDAYS, fetch_xlsx_spreadsheet, get_sheet_gids, nearest_weekday, sanitize_sheet_name

from .schemas import Lesson


def use(
    processed_column: pd.Series,
    target: Target,
) -> Generator[CoreCourseEvent, None, None]:
    """
    Generate events from processed cells

    :param processed_column: series with processed cells (CoreCourseCell),
        multiindex with (weekday, timeslot) and (course, group) as name
    :param target: target to generate events for (needed for start and end dates)
    :param sheet_gids: sheet name -> gid mapping
    :return: generator of events
    """
    # -------- Iterate over processed cells --------
    (course, group) = processed_column.name
    course: str
    group: str

    for (weekday, timeslot), cell in processed_column.items():
        cell: CoreCourseCell | None
        if cell is None:
            continue
        weekday: str
        timeslot: tuple[datetime.time, datetime.time]

        event = convert_cell_to_event(
            cell=cell,
            weekday=weekday,
            timeslot=timeslot,
            course=course,
            group=group,
            target=target,
        )

        if event is None:
            continue

        yield event


async def get_all_lessons(parser_config: CoreCoursesConfig) -> list[Lesson]:
    parser = CoreCoursesParser()
    xlsx_file = await fetch_xlsx_spreadsheet(spreadsheet_id=parser_config.spreadsheet_id)
    original_target_sheet_names = [target.sheet_name for target in parser_config.targets]
    sheet_gids = await get_sheet_gids(parser_config.spreadsheet_id)
    pipeline_result = parser.pipeline(xlsx_file, original_target_sheet_names, sheet_gids, parser_config.spreadsheet_id)
    dfs_merged_ranges = parser.last_dfs_merged_ranges
    assert dfs_merged_ranges is not None

    all_lessons: list[Lesson] = []
    for target, grouped_dfs_with_cells_list in zip(parser_config.targets, pipeline_result):
        merged_ranges = dfs_merged_ranges.get(sanitize_sheet_name(target.sheet_name))

        # merge range index -> list of events
        merged_registry_for_events: dict[int, list[CoreCourseEvent]] = defaultdict(list)
        non_merged_events: list[CoreCourseEvent] = []

        for grouped_dfs_with_cells in grouped_dfs_with_cells_list:
            series_with_generators = grouped_dfs_with_cells.apply(use, target=target, sheet_gids=sheet_gids)
            for generator in series_with_generators:
                generator: Generator[CoreCourseEvent, None, None]

                for cell_event in generator:
                    if cell_event.subject in parser_config.ignored_subjects:
                        logger.debug(f"> Ignoring {cell_event.subject}")
                        continue

                    if merged_ranges:
                        if not cell_event.a1:
                            non_merged_events.append(cell_event)
                            continue
                        cell_row, cell_col = coordinate_to_tuple(cell_event.a1)
                        cell_row -= 1
                        cell_col -= 1
                        for i, (min_row, min_col, max_row, max_col) in enumerate(merged_ranges):
                            if (min_row <= cell_row <= max_row) and (min_col <= cell_col <= max_col):
                                merged_registry_for_events[i].append(cell_event)
                                break
                        else:
                            non_merged_events.append(cell_event)
                    else:
                        non_merged_events.append(cell_event)

        lessons_from_non_merged: list[Lesson] = []
        for cell_event in non_merged_events:
            if cell_event.location_item is None:
                lessons_from_non_merged.append(_event_to_lesson(cell_event))
            else:
                lessons_from_non_merged.extend(_process_location_item(cell_event, target))

        lessons_from_merged: list[Lesson] = []
        for merged_range_index, _events in merged_registry_for_events.items():
            assert merged_ranges is not None
            _events: list[CoreCourseEvent]
            groups = []
            merged_students_number = []
            merged_range = merged_ranges[merged_range_index]
            min_row, min_col, max_row, max_col = merged_range
            merged_range_a1_range = (
                f"{get_column_letter(min_col + 1)}{min_row + 1}:{get_column_letter(max_col + 1)}{max_row + 1}"
            )
            excel_ranges = []
            cell_event = _events[0]

            for cell_event in _events:
                if cell_event.group:
                    if cell_event.group not in groups and cell_event.group_student_number:
                        merged_students_number.append(cell_event.group_student_number)
                    if cell_event.group not in groups and cell_event.a1:
                        excel_ranges.append(cell_event.a1)
                else:
                    if cell_event.group_student_number:
                        merged_students_number.append(cell_event.group_student_number)
                    if cell_event.a1:
                        excel_ranges.append(cell_event.a1)

                if isinstance(cell_event.group, list | tuple):
                    groups.extend(cell_event.group)
                elif cell_event.group:
                    groups.append(cell_event.group)

            logger.info(groups)
            logger.info(merged_range_a1_range)
            logger.info(excel_ranges)
            logger.info(merged_students_number)

            if cell_event.location_item is None:
                lessons_from_merged.append(
                    _event_to_lesson(
                        cell_event,
                        group_name=tuple(groups),
                        students_number=sum(merged_students_number) if merged_students_number else None,
                        a1_range=merged_range_a1_range,
                    )
                )
            else:
                lessons_from_merged.extend(
                    _process_location_item(
                        cell_event,
                        target,
                        group_name=tuple(groups),
                        students_number=sum(merged_students_number) if merged_students_number else None,
                        a1_range=merged_range_a1_range,
                    )
                )
        logger.info(
            f"For {target.sheet_name} found {len(lessons_from_non_merged)} non-merged lessons and {len(lessons_from_merged)} merged (same a1 range) lessons, totaling {len(lessons_from_merged + lessons_from_non_merged)} lessons"
        )
        merged_lessons = merge_identical_lessons(lessons_from_merged + lessons_from_non_merged)
        logger.info(f"After merging identical lessons, for {target.sheet_name} found {len(merged_lessons)} lessons")
        all_lessons.extend(merged_lessons)

    all_lessons.sort(key=lambda x: (x.course_name, x.group_name, x.weekday, x.start_time))
    return all_lessons


def _event_to_lesson(
    cell_event: CoreCourseEvent,
    *,
    group_name: str | tuple[str, ...] | None = None,
    students_number: int | None = None,
    a1_range: str | None = None,
) -> Lesson:
    return Lesson(
        lesson_name=cell_event.subject,
        weekday=WEEKDAYS[cell_event.weekday],
        start_time=cell_event.start_time,
        end_time=cell_event.end_time,
        course_name=cell_event.course,
        group_name=group_name if group_name is not None else cell_event.group,
        teacher=cell_event.teacher,
        room=cell_event.location,
        date_on=None,
        date_except=None,
        students_number=students_number if students_number is not None else cell_event.group_student_number,
        spreadsheet_id=cell_event.spreadsheet_id,
        google_sheet_gid=cell_event.google_sheet_gid,
        google_sheet_name=cell_event.google_sheet_name,
        a1_range=a1_range if a1_range is not None else cell_event.a1,
    )


def _process_location_item(
    cell_event: CoreCourseEvent,
    target: Target,
    *,
    group_name: str | tuple[str, ...] | None = None,
    students_number: int | None = None,
    a1_range: str | None = None,
) -> list[Lesson]:
    location_item = cell_event.location_item
    assert location_item is not None
    starts = location_item.starts_from or target.start_date

    def convert_weeks_on_to_only_on(item: Item):
        if item.on_weeks:
            on = []
            for week in item.on_weeks:
                on_date = nearest_weekday(starts, cell_event.weekday) + datetime.timedelta(weeks=week - 1)
                on.append(on_date)
            if item.on:
                item.on.extend(on)
            elif on:
                item.on = on
        if item.on:
            item.on = sorted(set(item.on))

    lesson_start_time = cell_event.start_time
    lesson_end_time = cell_event.end_time

    if location_item.starts_at:
        _start_time = datetime.datetime.combine(starts, cell_event.start_time)
        _end_time = datetime.datetime.combine(starts, cell_event.end_time)
        duration = _end_time - _start_time
        lesson_start_time = location_item.starts_at
        lesson_end_time = (datetime.datetime.combine(starts, lesson_start_time) + duration).time()
    if location_item.till:
        lesson_end_time = location_item.till

    convert_weeks_on_to_only_on(location_item)

    main_lesson = Lesson(
        lesson_name=cell_event.subject,
        weekday=WEEKDAYS[cell_event.weekday],
        start_time=lesson_start_time,
        end_time=lesson_end_time,
        course_name=cell_event.course,
        group_name=group_name if group_name is not None else cell_event.group,
        teacher=cell_event.teacher,
        room=location_item.location or cell_event.location,
        date_on=location_item.on,
        date_except=location_item.except_,
        date_from=location_item.starts_from,
        students_number=students_number if students_number is not None else cell_event.group_student_number,
        spreadsheet_id=cell_event.spreadsheet_id,
        google_sheet_gid=cell_event.google_sheet_gid,
        google_sheet_name=cell_event.google_sheet_name,
        a1_range=a1_range if a1_range is not None else cell_event.a1,
    )

    lessons: list[Lesson] = [main_lesson]

    nested_on: list[Item] = []
    extra_nested: list[Item] = []
    if location_item.NEST:
        for item in location_item.NEST:
            convert_weeks_on_to_only_on(item)
            if item.on:
                nested_on.append(item)
            else:
                logger.info(f"Root Item: {location_item}, {item}")
                extra_nested.append(item)

    if extra_nested:  # TODO: Handle '421 (316 FROM 31/10)' case
        logger.warning(f"Extra nested is not implemented yet\nItem({location_item})")

    for item in nested_on:
        assert item.on, f"Item {item} has no on"
        if item.location:
            main_lesson.date_except = (main_lesson.date_except or []) + item.on

        nested_lesson = main_lesson.model_copy()
        nested_lesson.date_on = item.on
        nested_lesson.room = item.location or main_lesson.room
        nested_lesson.start_time = item.starts_at or main_lesson.start_time
        nested_lesson.end_time = item.till or main_lesson.end_time
        lessons.append(nested_lesson)

    return lessons


def _are_lessons_identical(lesson1: Lesson, lesson2: Lesson) -> bool:
    """Check if two lessons are identical (excluding Excel cell location)"""
    return (
        lesson1.lesson_name == lesson2.lesson_name
        and lesson1.weekday == lesson2.weekday
        and lesson1.start_time == lesson2.start_time
        and lesson1.end_time == lesson2.end_time
        and lesson1.room == lesson2.room
        and lesson1.teacher == lesson2.teacher
        and lesson1.date_on == lesson2.date_on
        and lesson1.date_except == lesson2.date_except
    )


def merge_identical_lessons(lessons: list[Lesson]) -> list[Lesson]:
    groups: list[list[Lesson]] = []
    for lesson in lessons:
        # Find if this lesson belongs to an existing group
        found_group = False
        for group in groups:
            if _are_lessons_identical(lesson, group[0]):
                group.append(lesson)
                found_group = True
                break
        if not found_group:
            groups.append([lesson])

    result = []
    for group in groups:
        if len(group) > 1:
            excel_ranges = [lesson.a1_range for lesson in group if lesson.a1_range is not None]
            merged_groups = []
            for lesson in group:
                if lesson.group_name is not None:
                    if isinstance(lesson.group_name, tuple):
                        merged_groups.extend(lesson.group_name)
                    else:
                        merged_groups.append(lesson.group_name)
            students_numbers = [lesson.students_number for lesson in group if lesson.students_number is not None]
            students_number = sum(students_numbers) if students_numbers else None
            lesson = group[0].model_copy()
            lesson.a1_range = ";".join(excel_ranges)
            lesson.group_name = tuple(sorted(merged_groups))
            lesson.students_number = students_number
            result.append(lesson)
        else:
            result.append(group[0])
    return result
